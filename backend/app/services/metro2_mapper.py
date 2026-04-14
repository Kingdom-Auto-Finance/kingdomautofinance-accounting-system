"""Metro 2 column mapping engine - Layer 1 of the guardrail system.

Backs the File Upload tab's Map Fields modal. Given an uploaded CSV/XLSX
file, parses the headers, suggests mappings (header → Metro 2 field) based
on name similarity, and persists operator-approved mappings as reusable
templates.

This module is pure-Python (no DB writes except for templates). The API
layer is responsible for staging raw_rows into metro2_upload_batches.
"""
from __future__ import annotations

import csv
import difflib
import io
import logging
import re
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

from app.db.supabase_client import get_supabase_client
from app.services import metro2_schema as schema

logger = logging.getLogger(__name__)

# Common header synonyms used by auto-suggest. Lower-case keys, Metro 2
# field-name values. Covers frequent Switch Labs / spreadsheet conventions.
SYNONYMS: Dict[str, str] = {
    "account_number": "ConsumerAccountNumber",
    "account number": "ConsumerAccountNumber",
    "acct_num": "ConsumerAccountNumber",
    "acctnum": "ConsumerAccountNumber",
    "acct#": "ConsumerAccountNumber",
    "loan_id": "ConsumerAccountNumber",
    "loan number": "ConsumerAccountNumber",

    "balance": "CurrentBalance",
    "current_balance": "CurrentBalance",
    "curr_bal": "CurrentBalance",
    "outstanding": "CurrentBalance",

    "past_due": "AmountPastDue",
    "amt_past_due": "AmountPastDue",
    "past due amount": "AmountPastDue",
    "delinquent": "AmountPastDue",

    "orig_loan": "HighestCreditOrOrigLoanAmt",
    "original_loan": "HighestCreditOrOrigLoanAmt",
    "original amount": "HighestCreditOrOrigLoanAmt",
    "loan_amount": "HighestCreditOrOrigLoanAmt",
    "loan amt": "HighestCreditOrOrigLoanAmt",

    "term": "TermsDuration",
    "term_months": "TermsDuration",
    "months": "TermsDuration",

    "freq": "TermsFrequency",
    "frequency": "TermsFrequency",
    "pay_freq": "TermsFrequency",

    "status": "AccountStatus",
    "account_status": "AccountStatus",
    "metro_status": "AccountStatus",

    "opened": "DateOpened",
    "date_opened": "DateOpened",
    "origination_date": "DateOpened",
    "contract_date": "DateOpened",

    "closed": "DateClosed",
    "date_closed": "DateClosed",

    "last_payment": "DateLastPayment",
    "last_pmt_date": "DateLastPayment",
    "lpd": "DateLastPayment",

    "dob": "DateOfBirth",
    "birth_date": "DateOfBirth",
    "date_of_birth": "DateOfBirth",

    "ssn": "SSN",
    "social": "SSN",
    "social_security": "SSN",

    "first": "FirstName",
    "first_name": "FirstName",
    "firstname": "FirstName",
    "fname": "FirstName",

    "last": "Surname",
    "last_name": "Surname",
    "lastname": "Surname",
    "surname": "Surname",
    "lname": "Surname",

    "middle": "MiddleName",
    "middle_name": "MiddleName",
    "mname": "MiddleName",

    "phone": "PhoneNumber",
    "phone_number": "PhoneNumber",
    "telephone": "PhoneNumber",
    "cell": "PhoneNumber",

    "address": "Address1",
    "address_1": "Address1",
    "addr1": "Address1",
    "street": "Address1",

    "address_2": "Address2",
    "addr2": "Address2",
    "unit": "Address2",
    "apt": "Address2",

    "city": "City",
    "state": "State",
    "zip": "PostalCode",
    "zipcode": "PostalCode",
    "postal_code": "PostalCode",

    "scheduled_payment": "ScheduledPaymentAmt",
    "monthly_payment": "ScheduledPaymentAmt",
    "payment_amt": "ScheduledPaymentAmt",

    "actual_payment": "ActualPaymentAmt",
    "last_payment_amt": "ActualPaymentAmt",

    "fcra": "FCRA_DOFI",
    "dofi": "FCRA_DOFI",
    "date_first_delinquency": "FCRA_DOFI",

    "chargeoff": "OriginalChargeoffAmt",
    "charge_off": "OriginalChargeoffAmt",
    "original_chargeoff": "OriginalChargeoffAmt",

    "ecoa": "ECOACode",
}


@dataclass
class MappingSuggestion:
    source_column: str
    suggested_field: Optional[str]
    confidence: float              # 0.0-1.0
    sample_values: List[str]
    importance: Optional[str]      # required/recommended/optional of suggested_field


# ─── Public API ──────────────────────────────────────────────────────────────
def parse_upload(file_bytes: bytes,
                 filename: str) -> Tuple[List[str], List[Dict[str, str]]]:
    """Parse an uploaded CSV or XLSX file into (headers, rows).

    XLSX support requires openpyxl. If the file is XLSX and openpyxl is not
    installed, raises ValueError with a clear message.
    """
    lower = filename.lower()
    if lower.endswith(".xlsx") or lower.endswith(".xls"):
        return _parse_xlsx(file_bytes)
    return _parse_csv(file_bytes)


def suggest_mappings(headers: List[str],
                     sample_rows: List[Dict[str, str]]) -> List[MappingSuggestion]:
    """For each upload header, suggest the best-matching Metro 2 field.

    Suggests via:
      1. Exact-match synonym lookup (confidence 1.0).
      2. Normalized (snake_case) substring match against field names (0.8).
      3. difflib closest-match against field names (0.6+ if ratio ≥ 0.7).
      4. None (operator must pick manually).
    """
    field_names = [f.name for f in schema.FIELDS]
    norm_field_names = {_normalize(n): n for n in field_names}

    suggestions: List[MappingSuggestion] = []
    for header in headers:
        norm = _normalize(header)
        suggested: Optional[str] = None
        confidence = 0.0

        if norm in SYNONYMS:
            suggested = SYNONYMS[norm]
            confidence = 1.0
        elif norm in norm_field_names:
            suggested = norm_field_names[norm]
            confidence = 0.95
        else:
            # Substring / contained match on normalized names.
            for nf, original in norm_field_names.items():
                if nf in norm or norm in nf:
                    suggested = original
                    confidence = 0.8
                    break
            if not suggested:
                close = difflib.get_close_matches(norm, list(norm_field_names.keys()),
                                                  n=1, cutoff=0.7)
                if close:
                    suggested = norm_field_names[close[0]]
                    confidence = difflib.SequenceMatcher(None, norm, close[0]).ratio()

        samples = _collect_samples(sample_rows, header, limit=3)
        importance = (
            schema.FIELDS_BY_NAME[suggested].importance if suggested else None
        )
        suggestions.append(MappingSuggestion(
            source_column=header,
            suggested_field=suggested,
            confidence=round(confidence, 2),
            sample_values=samples,
            importance=importance,
        ))

    return suggestions


def validate_mapping(mapping: Dict[str, str]) -> Dict[str, Any]:
    """Check that a proposed mapping covers all required Metro 2 fields.

    Returns a dict matching the shape Switch Labs' counter expects:
      {
        "total_required": int,
        "mapped_required": int,
        "missing_required": [field_name, ...],
        "total_recommended": int,
        "mapped_recommended": int,
        "missing_recommended": [...],
        "is_valid": bool  # True iff every required field is mapped
      }
    """
    mapped_fields = set(mapping.values())
    required = {f.name for f in schema.REQUIRED_FIELDS}
    recommended = {f.name for f in schema.RECOMMENDED_FIELDS}

    missing_req = sorted(required - mapped_fields)
    missing_rec = sorted(recommended - mapped_fields)

    return {
        "total_required": len(required),
        "mapped_required": len(required - set(missing_req)),
        "missing_required": missing_req,
        "total_recommended": len(recommended),
        "mapped_recommended": len(recommended - set(missing_rec)),
        "missing_recommended": missing_rec,
        "is_valid": not missing_req,
    }


def apply_mapping(rows: List[Dict[str, str]],
                  mapping: Dict[str, str]) -> List[Dict[str, Any]]:
    """Transform raw CSV rows into Metro 2-keyed record dicts via a mapping."""
    out: List[Dict[str, Any]] = []
    for r in rows:
        rec: Dict[str, Any] = {}
        for source_col, target_field in mapping.items():
            if source_col in r:
                rec[target_field] = r[source_col]
        out.append(rec)
    return out


# ─── Template persistence ────────────────────────────────────────────────────
def list_templates() -> List[Dict[str, Any]]:
    sb = get_supabase_client()
    res = sb.table("metro2_mapping_templates").select("*").order(
        "is_default", desc=True
    ).order("updated_at", desc=True).execute()
    return res.data or []


def save_template(name: str, mapping: Dict[str, str],
                  description: Optional[str] = None,
                  is_default: bool = False,
                  user_id: Optional[str] = None) -> Dict[str, Any]:
    sb = get_supabase_client()
    if is_default:
        sb.table("metro2_mapping_templates").update({
            "is_default": False
        }).neq("id", "00000000-0000-0000-0000-000000000000").execute()

    # Upsert-by-name.
    existing = sb.table("metro2_mapping_templates").select("id").eq(
        "name", name
    ).execute().data
    if existing:
        res = sb.table("metro2_mapping_templates").update({
            "mapping": mapping,
            "description": description,
            "is_default": is_default,
        }).eq("id", existing[0]["id"]).execute()
        return res.data[0] if res.data else {}

    res = sb.table("metro2_mapping_templates").insert({
        "name": name,
        "description": description,
        "mapping": mapping,
        "is_default": is_default,
        "created_by": user_id,
    }).execute()
    return res.data[0] if res.data else {}


def delete_template(template_id: str) -> None:
    sb = get_supabase_client()
    sb.table("metro2_mapping_templates").delete().eq("id", template_id).execute()


# ─── Internal parsing helpers ────────────────────────────────────────────────
def _parse_csv(file_bytes: bytes) -> Tuple[List[str], List[Dict[str, str]]]:
    text = file_bytes.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    headers = list(reader.fieldnames or [])
    rows = [dict(r) for r in reader]
    return headers, rows


def _parse_xlsx(file_bytes: bytes) -> Tuple[List[str], List[Dict[str, str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise ValueError(
            "XLSX upload requires openpyxl; add to requirements.txt or "
            "upload as CSV instead."
        ) from e

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = [str(h) if h is not None else "" for h in next(rows_iter)]
    except StopIteration:
        return [], []
    rows: List[Dict[str, str]] = []
    for raw in rows_iter:
        rec = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            val = raw[i] if i < len(raw) else None
            rec[h] = "" if val is None else str(val)
        rows.append(rec)
    return headers, rows


def _normalize(s: str) -> str:
    """Normalize a column header for matching (lowercase, alnum+underscores)."""
    s = s.strip().lower()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    return s.strip("_")


def _collect_samples(rows: List[Dict[str, str]], header: str,
                     limit: int = 3) -> List[str]:
    out: List[str] = []
    for r in rows:
        val = r.get(header)
        if val in (None, ""):
            continue
        s = str(val).strip()
        if s and s not in out:
            out.append(s)
        if len(out) >= limit:
            break
    return out
